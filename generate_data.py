"""
Data processing script: reads the Excel workbook (318 door information) and
extracts embedded images into public/assets/catalog/, then writes src/data.ts.

Cabinet Door sheet (row 2 = header, data from row 3):
  - Rows above the \"Interior DooR\" divider = 柜门 (frameCategory cabinet).
  - Rows below = 房门 (frameCategory room).
  - Column C = main door photo (list thumbnail); D = side view; E = profile section (UI: extra preview buttons).
  - Column R (18) = hinge picture for Step 6 (separate from door photos).
  - Profile / Cover|Insert from column F; hinge codes from column P (matched to hardware).

Rows using Excel DISPIMG() without true embedded drawings will have no extracted photos;
paste images into the sheet or add files under public/assets/catalog/cabinet-door/ manually.

Override path: set env GAINER_EXCEL to the .xlsx file.
"""

import os
import re
import shutil
import textwrap
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Resolve workbook path
# ---------------------------------------------------------------------------

def _resolve_xlsx() -> Path:
    env = os.environ.get("GAINER_EXCEL")
    if env:
        p = Path(env).expanduser().resolve()
        if p.is_file():
            return p
        raise FileNotFoundError(f"GAINER_EXCEL is not a file: {p}")
    here = Path(__file__).resolve().parent
    candidates = [
        here / "318 door information(3).xlsx",
        Path.home() / "Desktop" / "318 door information(3).xlsx",
    ]
    for p in candidates:
        if p.is_file():
            return p
    raise FileNotFoundError(
        "Place '318 door information(3).xlsx' next to generate_data.py, your Desktop, "
        "or set GAINER_EXCEL to the full path."
    )


XLSX_PATH = _resolve_xlsx()
PROJECT_ROOT = Path(__file__).resolve().parent
CATALOG_DIR = PROJECT_ROOT / "public" / "assets" / "catalog"


def _slug(s) -> str:
    s = str(s).strip() if s is not None else "x"
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", s)[:100] or "x"


def _best_image_for_row(ws, excel_row: int, prefer_cols: list[int]):
    """Pick one image anchored on excel_row; prefer columns in order, else smallest col."""
    found = []
    for img in ws._images or []:
        fr = img.anchor._from
        if fr.row + 1 != excel_row:
            continue
        col = fr.col + 1
        found.append((col, img))
    if not found:
        return None
    for pc in prefer_cols:
        for col, img in found:
            if col == pc:
                return img
    found.sort(key=lambda x: x[0])
    return found[0][1]


def _best_image_near_row(ws, excel_row: int, prefer_cols: list[int], row_slack: int):
    """Like _best_image_for_row but allow anchor row within ±row_slack (Excel paste offset)."""
    found = []
    for img in ws._images or []:
        fr = img.anchor._from
        er = fr.row + 1
        if abs(er - excel_row) > row_slack:
            continue
        col = fr.col + 1
        found.append((abs(er - excel_row), col, img))
    if not found:
        return None
    found.sort(key=lambda x: (x[0], x[1]))
    best_d = found[0][0]
    tier = [x for x in found if x[0] == best_d]
    for pc in prefer_cols:
        for _d, col, im in tier:
            if col == pc:
                return im
    tier.sort(key=lambda x: x[1])
    return tier[0][2]


def _image_at_row_col(ws, excel_row: int, col_1based: int):
    """Single embedded image anchored exactly on (row, col)."""
    for img in ws._images or []:
        fr = img.anchor._from
        if fr.row + 1 == excel_row and fr.col + 1 == col_1based:
            return img
    return None


def _save_image_at_cell(ws, excel_row: int, col_1based: int, subdir: str, basename: str):
    """Save the first embedded image anchored on (row, col), or None."""
    im = _image_at_row_col(ws, excel_row, col_1based)
    if not im:
        return None
    return _save_catalog_image(im, subdir, basename)


def _existing_catalog_url(subdir: str, basename: str) -> str | None:
    """If a file was already generated (or added manually), reuse its public URL."""
    safe = _slug(basename)
    out_dir = CATALOG_DIR / subdir
    for ext in ("png", "jpg", "jpeg", "webp"):
        p = out_dir / f"{safe}.{ext}"
        if p.is_file():
            return f"/assets/catalog/{subdir}/{safe}.{ext}"
    return None


def _extract_cabinet_door_image(
    ws,
    excel_row: int,
    col_1based: int,
    subdir: str,
    basename: str,
    *,
    prefer_cols: list[int],
    row_slack: int = 0,
) -> str | None:
    """
    Extract embedded image for Cabinet Door sheet cells.
    Tries exact (row,col), then same-row preferred columns, then ±row_slack row;
    finally reuses an existing file on disk (e.g. last good export).
    """
    order: list[int] = []
    for c in [col_1based, *prefer_cols]:
        if c not in order:
            order.append(c)

    im = _image_at_row_col(ws, excel_row, col_1based)
    if not im:
        im = _best_image_for_row(ws, excel_row, order)
    if not im and row_slack > 0:
        im = _best_image_near_row(ws, excel_row, order, row_slack=row_slack)
    if im:
        url = _try_save_catalog_image(im, subdir, basename)
        if url:
            return url
    return _existing_catalog_url(subdir, basename)


def _worksheet(wb, *names: str):
    """Resolve first existing sheet by case-insensitive name."""
    lower_map = {n.lower(): n for n in wb.sheetnames}
    for n in names:
        key = n.lower()
        if key in lower_map:
            return wb[lower_map[key]]
    raise KeyError(f"No sheet among {names!r} in {wb.sheetnames!r}")


def _worksheet_optional(wb, *names: str):
    """Same as _worksheet but returns None if the sheet is absent."""
    try:
        return _worksheet(wb, *names)
    except KeyError:
        return None


def _save_catalog_image(img, subdir: str, basename: str) -> str:
    ext = (img.format or "png").lower()
    if ext == "jpeg":
        ext = "jpg"
    safe = _slug(basename)
    out_dir = CATALOG_DIR / subdir
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{safe}.{ext}"
    out_path.write_bytes(img._data())
    return f"/assets/catalog/{subdir}/{safe}.{ext}"


def _try_save_catalog_image(img, subdir: str, basename: str) -> str | None:
    """Some Excel DISPIMG / linked drawings raise on read; skip and fall back."""
    try:
        return _save_catalog_image(img, subdir, basename)
    except (ValueError, OSError, TypeError):
        return None


def _extract_row_images(ws, subdir: str, code_col_1based: int, prefer_cols: list[int], min_row: int = 2):
    """Map product code (str) -> public URL. Skips broken DISPIMG reads; keeps prior file on disk if any."""
    mapping = {}
    for r in range(min_row, ws.max_row + 1):
        code = ws.cell(r, code_col_1based).value
        if code is None or str(code).strip() in ("", "\\"):
            continue
        key = str(code).strip()
        im = _best_image_for_row(ws, r, prefer_cols)
        url = None
        if im:
            url = _try_save_catalog_image(im, subdir, key)
        if not url:
            url = _existing_catalog_url(subdir, key)
        if url:
            mapping[key] = url
    return mapping


def _extract_color_sheet(ws, subdir: str, min_row: int = 2):
    """Anodize / spray sheets: code col B, picture col C; null code uses slug(name)."""
    mapping = {}
    for r in range(min_row, ws.max_row + 1):
        name = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        if not name and not code:
            continue
        im = _best_image_for_row(ws, r, [3, 1, 2])
        key = str(code).strip() if code else _slug(name)
        url = None
        if im:
            url = _try_save_catalog_image(im, subdir, key)
        if not url:
            url = _existing_catalog_url(subdir, key)
        if not url:
            continue
        if code:
            mapping[str(code).strip()] = url
        if name:
            mapping[_slug(name)] = url
    return mapping


def _copy_catalog_asset(subdir: str, src_key: str, dst_key: str) -> str | None:
    """Copy an existing catalog file (e.g. reuse G31 photo for G32 when row has no drawing)."""
    out_dir = CATALOG_DIR / subdir
    src_safe = _slug(src_key)
    dst_safe = _slug(dst_key)
    for ext in ("png", "jpg", "jpeg", "webp"):
        sp = out_dir / f"{src_safe}.{ext}"
        if not sp.is_file():
            continue
        dp = out_dir / f"{dst_safe}.{ext}"
        shutil.copy2(sp, dp)
        return f"/assets/catalog/{subdir}/{dst_safe}.{ext}"
    return None


# ---------------------------------------------------------------------------
# Load workbook + build raw row dict (same shape as former raw_data.json)
# ---------------------------------------------------------------------------

wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)


def _sheet_rows(name: str):
    ws = wb[name]
    return [list(row) for row in ws.iter_rows(values_only=True)]


print(f"Loading: {XLSX_PATH}")
CATALOG_DIR.mkdir(parents=True, exist_ok=True)

_ws_glass = _worksheet(wb, "glass")
_ws_leather = _worksheet(wb, "leather")
_ws_wood = _worksheet_optional(wb, "wood veneer")
_ws_quartz = _worksheet(wb, "quartz stone")
_ws_hw = _worksheet(wb, "hardware")
_ws_handle = _worksheet_optional(wb, "handle")
_ws_cab = _worksheet(wb, "Cabinet Door", "cabinet door")

glass_pics = _extract_row_images(_ws_glass, "glass", 2, [3, 2])
# Row has no embedded image anchored on that line (e.g. G32) — reuse closest SKU photo until sheet is fixed.
for _gt, _gs in (("G32", "G31"),):  # golden tinted coated ≈ golden tinted
    if _gt not in glass_pics:
        _u = _copy_catalog_asset("glass", _gs, _gt)
        if _u:
            glass_pics[_gt] = _u
leather_pics = _extract_row_images(_ws_leather, "leather", 1, [3, 2])
wood_pics = (
    _extract_row_images(_ws_wood, "wood-veneer", 2, [1, 2]) if _ws_wood is not None else {}
)
quartz_pics = _extract_row_images(_ws_quartz, "quartz", 3, [2, 3])
anod_pics = _extract_color_sheet(_worksheet(wb, "Anodized Color"), "anodize")
soft_pics = _extract_color_sheet(_worksheet(wb, "Sprayed Soft-Touch Color"), "spray-soft")
metal_pics = _extract_color_sheet(_worksheet(wb, "Sprayed Metallic Color"), "spray-metallic")
hardware_pics = _extract_row_images(_ws_hw, "hardware", 4, [2, 1, 3])

# Also extract images for hardware rows that have no code, keyed by slugified name
def _extract_hardware_name_images(ws, subdir, prefer_cols, min_row=2):
    mapping = {}
    for r in range(min_row, ws.max_row + 1):
        code = ws.cell(r, 4).value
        name = ws.cell(r, 1).value
        if code or not name:
            continue
        key = _slug(str(name).replace("\n", " ").strip())
        im = _best_image_for_row(ws, r, prefer_cols)
        if not im:
            continue
        mapping[key] = _save_catalog_image(im, subdir, key)
    return mapping

hardware_name_pics = _extract_hardware_name_images(_ws_hw, "hardware", [2, 1, 3])
handle_pics = (
    _extract_row_images(_ws_handle, "handle", 1, [2, 3]) if _ws_handle is not None else {}
)

raw = {name: _sheet_rows(name) for name in wb.sheetnames}

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def parse_thickness(t: str):
    """'3/4/5/8' -> [3,4,5,8]  ;  '4+4/5+5' -> [8,10]  ;  '4+4/4+5/5+5' -> [8,9,10]"""
    if t is None or t is False:
        return []
    t = str(t)
    if not t or t.strip() in ("\\", "", "None"):
        return []
    parts = [p.strip() for p in t.replace("mm", "").split("/") if p.strip() and p.strip() != "\\"]
    result = []
    for p in parts:
        if "+" in p:
            nums = [float(x) for x in p.split("+")]
            total = sum(nums)
            result.append(int(total) if total == int(total) else total)
        else:
            try:
                val = float(p)
                result.append(int(val) if val == int(val) else val)
            except ValueError:
                pass
    return sorted(set(result))


def parse_size_limits(notes: str):
    """Extract minW, maxW, minH, maxH from the Notes field."""
    if not notes:
        return {"minW": None, "maxW": None, "minH": None, "maxH": None}
    minW = maxW = minH = maxH = None
    nf = notes.replace("\n", " ").replace("：", ":").replace("≤", "<=").replace("≥", ">=")
    nf = nf.replace("＜", "<").replace("＞", ">")

    # --- Height ---
    # "Door height <= 2700 mm"
    m = re.search(r'(?:Door\s+)?height\s*:?\s*(?:H\s*)?<=?\s*(\d+)', nf, re.I)
    if m:
        maxH = int(m.group(1))
    # standalone "H <= 2700" (word boundary to avoid matching "width")
    m = re.search(r'(?<![a-zA-Z])H\s*<=?\s*(\d+)', nf)
    if m:
        maxH = int(m.group(1))

    # --- Width ---
    # "Door width <= 500 mm"
    m = re.search(r'(?:Door\s+)?width\s*:?\s*(?:W\s*)?<=?\s*(\d+)', nf, re.I)
    if m:
        maxW = int(m.group(1))
    # standalone "W <= 500"  (word boundary)
    m = re.search(r'(?<![a-zA-Z])W\s*<=?\s*(\d+)', nf)
    if m:
        maxW = int(m.group(1))

    # 350mm < W < 600mm  /  350＜W＜600
    m = re.search(r'(\d+)\s*(?:mm)?\s*<\s*W\s*<\s*(\d+)', nf, re.I)
    if m:
        minW = int(m.group(1))
        maxW = int(m.group(2))

    # 1200 <= W <= 2400
    m = re.search(r'(\d+)\s*(?:mm)?\s*<=?\s*W\s*<=?\s*(\d+)', nf, re.I)
    if m:
        minW = int(m.group(1))
        maxW = int(m.group(2))

    # 1800 <= H <= 3000
    m = re.search(r'(\d+)\s*(?:mm)?\s*<=?\s*(?<![a-zA-Z])H\s*<=?\s*(\d+)', nf)
    if m:
        minH = int(m.group(1))
        maxH = int(m.group(2))

    return {"minW": minW, "maxW": maxW, "minH": minH, "maxH": maxH}


def parse_door_thickness(notes: str):
    if not notes:
        return None
    notes_flat = notes.replace("\n", " ").replace("：", ":")
    m = re.search(r'(?:door|Door)\s*thickness\s*:\s*([\d.]+)', notes_flat, re.I)
    if m:
        return float(m.group(1))
    m = re.search(r'door\s*thickness\s*[：:]\s*([\d.]+)', notes_flat, re.I)
    if m:
        return float(m.group(1))
    m = re.search(r'D\s*=\s*([\d.]+)', notes_flat)
    if m:
        return float(m.group(1))
    return None


def parse_frame_profile_mounting(cell):
    """Excel 'Thinkness' column, e.g. '28mm/Cover', '22mm/insert', '32mm/Aluminum Stripe'."""
    if not cell or str(cell).strip() in ("\\", ""):
        return None, None
    label = str(cell).replace("\n", " ").strip()
    low = label.lower()
    mounting = None
    if "insert" in low:
        mounting = "insert"
    elif "cover" in low:
        mounting = "cover"
    return label, mounting


def parse_glass_prices_cell(text):
    """Parse 'normal glass:680/m²' style lines from price column."""
    if not text:
        return None
    t = str(text).replace("\n", " ")

    def grab(pat):
        m = re.search(pat, t, re.I)
        return int(m.group(1)) if m else None

    out = {
        "normalGlass": grab(r"normal\s*glass\s*[：:]\s*(\d+)"),
        "blackGlass": grab(r"black\s*glass\s*[：:]\s*(\d+)"),
        "coatedGlass": grab(r"coated\s*glass\s*[：:]\s*(\d+)"),
    }
    if out["normalGlass"] is None and out["blackGlass"] is None and out["coatedGlass"] is None:
        return None
    return out


def parse_hinge_codes(raw):
    if not raw or str(raw).strip() in ("\\", "/", "", "None"):
        return []
    parts = re.split(r"[/\n,]+", str(raw))
    return [p.strip() for p in parts if p.strip() and p.strip() not in ("/", "\\")]


def _hinge_code_fingerprint(s: str) -> str:
    """Lowercase alphanumeric only — matches AN1723D to AN 172 3D."""
    return re.sub(r"[^a-z0-9]", "", str(s).strip().lower())


# Cabinet Door column P typos / variants → canonical code string on hardware sheet
_HINGE_CODE_CANONICAL_BY_FINGERPRINT = {
    "an1723d": "AN 172 3D",
}


def normalize_hinge_code_for_catalog(token: str) -> str:
    """Map Excel hinge tokens to hardware sheet codes where we know aliases."""
    t = str(token).strip()
    if not t:
        return t
    fp = _hinge_code_fingerprint(t)
    return _HINGE_CODE_CANONICAL_BY_FINGERPRINT.get(fp, t)


def infer_specific_colors_from_filler(std_filler_raw: str):
    if not std_filler_raw or str(std_filler_raw).strip() in ("\\", "color swatch colors"):
        return None
    text = str(std_filler_raw)
    if "swatch" in text.lower():
        return None
    codes = re.findall(r"\b[A-Z]{2,3}\d{2,}[A-Z0-9]*\b", text)
    return list(dict.fromkeys(codes)) if codes else None


def classify_door_type(dt: str):
    if not dt:
        return "other"
    dt_low = dt.lower().replace("\n", " ")
    if "quartz" in dt_low and "glass" in dt_low:
        return "mixed"
    if "quartz" in dt_low:
        return "aluminum frame quartz stone door"
    if "leather" in dt_low or "fiber" in dt_low:
        return "aluminum frame fiber/leather door"
    if "wood" in dt_low:
        return "aluminum frame wood door"
    if "pivot" in dt_low:
        return "aluminum frame glass pivot door"
    if "slide" in dt_low:
        return "aluminum frame glass slide door"
    if "glass" in dt_low:
        return "aluminum frame glass door"
    return dt.replace("\n", " ").strip()


def infer_allowed_fillers(door_type_raw: str, std_filler: str):
    fillers = set()
    dt = (door_type_raw or "").lower().replace("\n", " ")
    sf = (std_filler or "").lower()
    if "glass" in dt or "G0" in (std_filler or "") or "G3" in (std_filler or ""):
        fillers.add("glass")
    if "quartz" in dt or "quartz" in sf:
        fillers.add("quartz stone")
    if "leather" in dt or "fiber" in dt:
        fillers.add("leather")
    if "wood" in dt:
        fillers.add("wood veneer")
    if not fillers:
        if std_filler and ("G01" in std_filler or "G33" in std_filler):
            fillers.add("glass")
    return sorted(fillers) if fillers else ["glass"]


def parse_allowed_finishing(surface_raw: str, color_raw: str):
    """Derive allowedFinishing list from raw surface/color fields."""
    if not surface_raw:
        return ["anodize", "spraySoftTouch", "sprayMetallic"]
    sf = surface_raw.lower().replace("\n", " ")
    result = []
    if "anod" in sf:
        result.append("anodize")
    if "spray" in sf:
        result.append("spraySoftTouch")
        result.append("sprayMetallic")
    if not result:
        result = ["anodize", "spraySoftTouch", "sprayMetallic"]
    return result


def parse_hardware_colors(c: str):
    if not c:
        return []
    return [x.strip().lower() for x in c.replace("/", ",").split(",") if x.strip()]


def _surface_pic(pic_map: dict, code, name):
    if code:
        u = pic_map.get(str(code).strip())
        if u:
            return u
    if name:
        return pic_map.get(_slug(str(name).replace("\n", " ").strip()))
    return None


def to_ts_val(v):
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, list):
        inner = ", ".join(to_ts_val(x) for x in v)
        return f"[{inner}]"
    s = str(v).replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n")
    return f"'{s}'"


# ---------------------------------------------------------------------------
# Cabinet Door — header row (row 2) column map (supports inserted handle picture cols)
# ---------------------------------------------------------------------------

def _cab_header_texts(ws, header_row: int = 2) -> dict[int, str]:
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        out[c] = str(v).strip().lower().replace("\n", " ") if v else ""
    return out


def _cab_first_col(texts: dict[int, str], pred) -> int | None:
    for c in sorted(texts.keys()):
        t = texts[c]
        if t and pred(t):
            return c
    return None


def _cabinet_door_column_map(ws, header_row: int = 2) -> dict:
    tmap = _cab_header_texts(ws, header_row)

    handle_pic_cols = []
    for c in sorted(tmap.keys()):
        s = tmap[c]
        if not s or "hinge" in s:
            continue
        if "handle" in s and "picture" in s:
            handle_pic_cols.append(c)
    handle_pic_cols = sorted(set(handle_pic_cols))

    handle_col = None
    scored = []
    for c in sorted(tmap.keys()):
        s = tmap[c]
        if not s or "hinge" in s:
            continue
        if "picture" in s and "handle" in s:
            continue
        if "diagram" in s or ("position" in s and "handle" in s):
            continue
        if "handle" not in s:
            continue
        score = 0
        if "type" in s:
            score += 3
        if s.strip() == "handle":
            score += 2
        scored.append((-score, c))
    scored.sort()
    if scored:
        handle_col = scored[0][1]
    else:
        handle_col = 7

    diagram_col = _cab_first_col(
        tmap, lambda s: "diagram" in s or ("position" in s and "handle" in s)
    )

    def col_or(*words, default: int) -> int:
        hit = _cab_first_col(tmap, lambda s: all(w in s for w in words))
        return hit if hit is not None else default

    hinge_code_c = col_or("hinge", "code", default=16)
    hw_color_c = None
    for c in range(hinge_code_c + 1, min(hinge_code_c + 6, ws.max_column + 1)):
        s = (tmap.get(c, "") or "").strip()
        if s == "color":
            hw_color_c = c
            break
    if hw_color_c is None:
        hw_color_c = 17

    # Merged header row often leaves handle thumbnail cols blank; they sit right of "handle Type".
    if not handle_pic_cols and handle_col is not None:
        cand = [handle_col + i for i in (1, 2, 3) if handle_col + i <= ws.max_column]
        images = getattr(ws, "_images", None) or []
        found = []
        for c in cand:
            for img in images:
                fr = img.anchor._from
                if fr.col + 1 != c or fr.row + 1 <= header_row:
                    continue
                found.append(c)
                break
        handle_pic_cols = sorted(set(found)) if found else cand

    return {
        "picture_main": 3,
        "picture_side": 4,
        "picture_profile": 5,
        "thinkness": 6,
        "handle": handle_col,
        "handle_pic_cols": handle_pic_cols,
        "diagram": diagram_col,
        "surface": col_or("surface", "finish", default=8),
        "door_type": col_or("door", "type", default=9),
        "std_filler": _cab_first_col(
            tmap,
            lambda s: "filler" in s and "standard" in s and "optional" not in s,
        )
        or 10,
        "price": _cab_first_col(tmap, lambda s: s == "price") or 11,
        "filler_opt": _cab_first_col(
            tmap, lambda s: "filler" in s and "optional" in s
        )
        or 12,
        "price_m2": _cab_first_col(tmap, lambda s: "price/m" in s or "m²" in s or "m2" in s)
        or 13,
        "filler_thickness": _cab_first_col(
            tmap,
            lambda s: "thickness" in s and "filler" in s,
        )
        or 14,
        "hardware": _cab_first_col(
            tmap, lambda s: "other hardware" in s or s == "other hardware"
        )
        or 15,
        "hinge_code": hinge_code_c,
        "hw_color": hw_color_c,
        "hinge_picture": _cab_first_col(
            tmap, lambda s: "hinge" in s and "picture" in s
        )
        or 18,
        "notes": _cab_first_col(tmap, lambda s: "note" in s) or 19,
    }


def _fix_thinkness_col(tmap: dict[int, str], default: int = 6) -> int:
    c = _cab_first_col(tmap, lambda s: "thinkness" in s)
    if c is not None:
        return c
    c2 = _cab_first_col(tmap, lambda s: "thickness" in s and "filler" not in s)
    return c2 if c2 is not None else default


def parse_handle_workflow(cell) -> str:
    if not cell or str(cell).strip() in ("", "\\", "None"):
        return "none"
    s = str(cell).strip().lower()
    if "without" in s and "handle" in s:
        return "none"
    if "no handle" in s:
        return "none"
    if "无拉手" in str(cell):
        return "none"
    if "v shape" in s or "v-shape" in s or "vshape" in s:
        return "vshape"
    if s.strip() == "handle":
        return "separate"
    if "separate" in s:
        return "separate"
    if ("fixed" in s or "款式固定" in str(cell)) and "separate" not in s:
        return "fixed"
    if "cnc" in s or "routed" in s or "铣" in str(cell):
        return "cnc"
    return "legacy_catalog"


def _parse_paren_codes(text: str) -> list[str]:
    return re.findall(r"\(([A-Za-z0-9][A-Za-z0-9._-]*)\)", str(text))


def _tokenize_handle_codes(raw: str) -> list[str]:
    parts = re.split(r"[/\n,，]+", str(raw))
    out = []
    for p in parts:
        p = p.strip()
        if len(p) < 2:
            continue
        pl = p.lower()
        if pl in (
            "separate handle",
            "routed handle(cnc)",
            "cnc",
            "handle",
            "fixed",
            "pull",
            "without handle",
        ):
            continue
        if "(" in p:
            out.extend(_parse_paren_codes(p))
        elif re.match(r"^[A-Za-z0-9][A-Za-z0-9._-]{1,}$", p):
            out.append(p)
    seen = set()
    uniq = []
    for x in out:
        k = x.upper()
        if k not in seen:
            seen.add(k)
            uniq.append(x)
    return uniq


_HANDLE_CODE_STOPWORDS = frozenset(
    {"fixed", "handle", "cnc", "separate", "pull", "none", "without", "shape"}
)


def _extract_row_handle_options(ws, row: int, pic_cols: list[int], frame_id: str) -> list[dict]:
    opts = []
    seen = set()
    door_key = str(ws.cell(row, 1).value or "").strip()
    for i, col in enumerate(pic_cols):
        basename = f"{frame_id}-handle-opt{i + 1}"
        url = _extract_cabinet_door_image(
            ws,
            row,
            col,
            "cabinet-door",
            basename,
            prefer_cols=[col, col + 1, col - 1],
            row_slack=0,
        )
        code = None
        # Codes sit in the same row as the thumbnails (cols 8–10) or in cells below;
        # do not read column left of the thumbnail (that is "handle Type" on other rows → false "CNC").
        for dr in range(0, 8):
            r2 = row + dr
            if dr > 0:
                c1 = ws.cell(r2, 1).value
                sk = str(c1).strip() if c1 else ""
                if sk and sk != door_key:
                    break
            for dc in (0, 1):
                cc = col + dc
                if cc < 1:
                    continue
                v = ws.cell(r2, cc).value
                if not v:
                    continue
                st = str(v).strip()
                m = re.search(r"\(([A-Za-z0-9][A-Za-z0-9._-]*)\)", st)
                if m:
                    code = m.group(1)
                    break
                m2 = re.search(
                    r"\b([A-Za-z]{1,4}\d{2,6}(?:-\d+)?)\b",
                    st,
                )
                if m2:
                    code = m2.group(1)
                    break
                if re.match(r"^[A-Za-z0-9][A-Za-z0-9._-]{2,}$", st):
                    code = st
                    break
            if code:
                break
        if code and str(code).strip().lower() in _HANDLE_CODE_STOPWORDS:
            code = None
        if not code:
            code = f"H{i + 1}"
        if code in seen:
            code = f"{code}-{i + 1}"
        seen.add(code)
        if url or code != f"H{i + 1}":
            opts.append({"code": code, "name": code, "picture": url})
    return opts


def _options_from_text_and_pics(raw: str, pics: dict) -> list[dict]:
    opts = []
    for code in _tokenize_handle_codes(raw):
        opts.append(
            {
                "code": code,
                "name": code,
                "picture": pics.get(code) or pics.get(str(code).strip()),
            }
        )
    return opts


def _infer_fixed_handle_code(wf: str, handle_raw, options: list[dict]) -> str | None:
    if wf == "vshape":
        if options:
            return options[0]["code"]
        toks = _tokenize_handle_codes(str(handle_raw or ""))
        return toks[0] if toks else None
    if wf == "fixed":
        if options:
            return options[0]["code"]
        toks = _tokenize_handle_codes(str(handle_raw or ""))
        junk = {"fixed", "handle", "pull", "cnc", "separate", "none"}
        for t in toks:
            if str(t).strip().lower() in junk:
                continue
            return t
        return "FIXED"
    return None


# ---------------------------------------------------------------------------
# 1. FRAMES — sheet "Cabinet Door": row 2 = header, row 3+ = data;
#    section divider cell "Interior DooR" starts 房门 (room) category.
# ---------------------------------------------------------------------------
ws_cab = _ws_cab
_cab_hdr = _cab_header_texts(ws_cab, 2)
_cab_cols = _cabinet_door_column_map(ws_cab, 2)
_cab_cols["thinkness"] = _fix_thinkness_col(_cab_hdr, 6)

frames = []
frame_standard_glass_prices = {}

category = "cabinet"
for r in range(3, ws_cab.max_row + 1):
    code_cell = ws_cab.cell(r, 1).value
    if code_cell is None or str(code_cell).strip() == "":
        continue
    head = str(code_cell).strip()
    if head.lower() == "cabinet door":
        continue
    if re.search(r"interior", head, re.I) and "door" in head.lower():
        category = "room"
        continue

    cc = _cab_cols
    door_type_raw = ws_cab.cell(r, cc["door_type"]).value
    if not door_type_raw or str(door_type_raw).strip() in ("", "\\"):
        continue

    code = str(code_cell).strip()
    frame_id = f"{category}-{code}"

    picture_main = _extract_cabinet_door_image(
        ws_cab,
        r,
        cc["picture_main"],
        "cabinet-door",
        f"{frame_id}-main",
        prefer_cols=[3, 4, 5, 2, 6],
        row_slack=1,
    )
    picture_side = _extract_cabinet_door_image(
        ws_cab,
        r,
        cc["picture_side"],
        "cabinet-door",
        f"{frame_id}-side",
        prefer_cols=[4, 3, 5],
    )
    picture_profile = _extract_cabinet_door_image(
        ws_cab,
        r,
        cc["picture_profile"],
        "cabinet-door",
        f"{frame_id}-profile",
        prefer_cols=[5, 4, 3],
    )
    hinge_picture = _extract_cabinet_door_image(
        ws_cab,
        r,
        cc["hinge_picture"],
        "cabinet-door",
        f"{frame_id}-hinge",
        prefer_cols=[cc["hinge_picture"], cc["hinge_picture"] - 1, cc["hw_color"]],
    )
    picture = picture_main or picture_side or picture_profile
    pics = [u for u in (picture_main, picture_side, picture_profile) if u]

    think_raw = ws_cab.cell(r, cc["thinkness"]).value
    frame_profile_label, mounting_type = parse_frame_profile_mounting(think_raw)

    handle_raw = ws_cab.cell(r, cc["handle"]).value
    surface_raw = ws_cab.cell(r, cc["surface"]).value
    std_filler_raw = ws_cab.cell(r, cc["std_filler"]).value
    price_raw = ws_cab.cell(r, cc["price"]).value
    thickness_raw = ws_cab.cell(r, cc["filler_thickness"]).value
    hardware_raw = ws_cab.cell(r, cc["hardware"]).value
    hinge_code_raw = ws_cab.cell(r, cc["hinge_code"]).value
    hw_color_raw = ws_cab.cell(r, cc["hw_color"]).value
    notes = ws_cab.cell(r, cc["notes"]).value

    handle_diagram_picture = None
    if cc.get("diagram"):
        handle_diagram_picture = _extract_cabinet_door_image(
            ws_cab,
            r,
            cc["diagram"],
            "cabinet-door",
            f"{frame_id}-handle-diagram",
            prefer_cols=[cc["diagram"], cc["diagram"] + 1, cc["diagram"] - 1],
            row_slack=0,
        )
    if not handle_diagram_picture:
        handle_diagram_picture = picture_main

    door_type = classify_door_type(door_type_raw)
    allowed_fillers = infer_allowed_fillers(door_type_raw, std_filler_raw)
    filler_thickness = parse_thickness(str(thickness_raw)) if thickness_raw else []
    size_limits = parse_size_limits(notes)
    door_thickness = parse_door_thickness(notes)
    allowed_finishing = parse_allowed_finishing(surface_raw, None)

    specific_colors = infer_specific_colors_from_filler(std_filler_raw)

    std_fillers = []
    if std_filler_raw and str(std_filler_raw).strip() not in ("\\", "color swatch colors"):
        std_fillers = [s.strip() for s in str(std_filler_raw).replace("\n", "/").split("/") if s.strip()]

    handle_workflow = parse_handle_workflow(handle_raw)
    handle_options = _extract_row_handle_options(
        ws_cab, r, cc["handle_pic_cols"], frame_id
    )

    if handle_workflow == "separate" and not handle_options and handle_raw:
        handle_options = _options_from_text_and_pics(str(handle_raw), handle_pics)

    fixed_handle_code = _infer_fixed_handle_code(
        handle_workflow, handle_raw, handle_options
    )

    matched_handle = None
    if handle_workflow == "legacy_catalog" and handle_raw:
        matched_handle = str(handle_raw).replace("\n", " ").strip()
    elif handle_workflow == "separate" and not handle_options and handle_raw:
        matched_handle = str(handle_raw).replace("\n", " ").strip()
        handle_workflow = "legacy_catalog"

    matched_hardware = str(hardware_raw).replace("\n", " ").strip() if hardware_raw else None
    hw_colors = parse_hardware_colors(hw_color_raw)
    hinge_codes = [
        normalize_hinge_code_for_catalog(c) for c in parse_hinge_codes(hinge_code_raw)
    ]

    gp = parse_glass_prices_cell(price_raw)
    if gp:
        frame_standard_glass_prices[code] = gp

    frames.append({
        "id": frame_id,
        "code": code,
        "frameCategory": category,
        "doorType": door_type,
        "allowedFillers": allowed_fillers,
        "standardFillers": std_fillers,
        "fillerThicknessLimit": filler_thickness,
        "allowedFinishing": allowed_finishing,
        "specificColors": specific_colors,
        "sizeLimits": size_limits,
        "doorThickness": door_thickness,
        "frameProfileLabel": frame_profile_label,
        "mountingType": mounting_type,
        "handleWorkflow": handle_workflow,
        "handleOptions": handle_options,
        "fixedHandleCode": fixed_handle_code,
        "handleDiagramPicture": handle_diagram_picture,
        "matchedHandle": matched_handle,
        "matchedHardware": matched_hardware,
        "hingeCodes": hinge_codes,
        "hardwareColors": hw_colors,
        "pictureSideView": picture_side,
        "pictureProfile": picture_profile,
        "hingePicture": hinge_picture,
        "pictures": pics,
        "picture": picture,
    })

# ---------------------------------------------------------------------------
# 2. GLASS
# ---------------------------------------------------------------------------
standard_glass = {"G01", "G33", "G36"}
glass_list = []
for row in raw["glass"][1:]:
    clean = [c for c in row if c is not None]
    if not clean:
        continue
    name = row[0]
    code = row[1]
    craft = row[3] if len(row) > 3 else None
    craft_code = row[4] if len(row) > 4 else None
    thickness_raw = row[5] if len(row) > 5 else None
    silk_screen = row[6] if len(row) > 6 else None

    if not code:
        continue
    thicknesses = parse_thickness(str(thickness_raw)) if thickness_raw else []
    pricing_type = "standard" if code in standard_glass else "custom"

    glass_list.append({
        "code": code,
        "name": name,
        "type": "glass",
        "craft": craft if craft and craft != "\\" else None,
        "craftCode": craft_code if craft_code and craft_code != "\\" else None,
        "thicknesses": thicknesses,
        "silkScreen": True if silk_screen and silk_screen.lower() == "yes" else False,
        "pricingType": pricing_type,
        "priceSqm": None,
        "picture": glass_pics.get(str(code).strip()),
    })

# ---------------------------------------------------------------------------
# 3. LEATHER
# ---------------------------------------------------------------------------
leather_list = []
for row in raw["leather"][1:]:
    clean = [c for c in row if c is not None]
    if not clean:
        continue
    code = row[0]
    thickness_raw = row[1]
    material_desc = row[4] if len(row) > 4 else row[3] if len(row) > 3 else None

    if not code:
        continue
    thicknesses = parse_thickness(str(thickness_raw)) if thickness_raw else []
    mat_name = ""
    if material_desc:
        m = re.search(r'Front:\s*(.+?)(?:\n|$)', material_desc)
        if m:
            mat_name = m.group(1).strip()

    leather_list.append({
        "code": code,
        "name": mat_name or code,
        "type": "leather",
        "thicknesses": thicknesses,
        "baseMaterial": "Honeycomb Aluminum",
        "pricingType": "custom",
        "picture": leather_pics.get(str(code).strip()),
    })

# ---------------------------------------------------------------------------
# 4. WOOD VENEER
# ---------------------------------------------------------------------------
wood_list = []
for row in raw.get("wood veneer", [])[1:]:
    clean = [c for c in row if c is not None]
    if not clean:
        continue
    code = row[1] if len(row) > 1 else None
    thickness_raw = row[2] if len(row) > 2 else None
    if not code:
        continue
    thicknesses = parse_thickness(str(thickness_raw)) if thickness_raw else []
    wood_list.append({
        "code": code,
        "name": code,
        "type": "woodVeneer",
        "thicknesses": thicknesses,
        "pricingType": "custom",
        "picture": wood_pics.get(str(code).strip()),
    })

# ---------------------------------------------------------------------------
# 5. QUARTZ STONE
# ---------------------------------------------------------------------------
quartz_list = []
for row in raw["quartz stone"][1:]:
    clean = [c for c in row if c is not None]
    if not clean:
        continue
    name = row[0]
    code = row[2] if len(row) > 2 else None
    surface = row[3] if len(row) > 3 else None
    thickness_raw = row[4] if len(row) > 4 else None
    if not code:
        continue
    thicknesses = parse_thickness(str(thickness_raw).replace("mm", "")) if thickness_raw else []
    quartz_list.append({
        "code": code,
        "name": name,
        "type": "quartzStone",
        "surface": surface,
        "thicknesses": thicknesses,
        "pricingType": "custom",
        "picture": quartz_pics.get(str(code).strip()),
    })

# ---------------------------------------------------------------------------
# 6. SURFACE FINISHES
# ---------------------------------------------------------------------------
anodize_colors = []
for row in raw["Anodized Color"][1:]:
    clean = [c for c in row if c is not None]
    if not clean:
        continue
    name = row[0]
    code = row[1] if len(row) > 1 else None
    if not name:
        continue
    nm = name.replace("\n", " ").strip()
    anodize_colors.append({
        "code": code,
        "name": nm,
        "picture": _surface_pic(anod_pics, code, nm),
    })

spray_soft_colors = []
for row in raw["Sprayed Soft-Touch Color"][1:]:
    clean = [c for c in row if c is not None]
    if not clean:
        continue
    name = row[0]
    code = row[1] if len(row) > 1 else None
    if not name:
        continue
    nm = name.replace("\n", " ").strip()
    spray_soft_colors.append({
        "code": code,
        "name": nm,
        "picture": _surface_pic(soft_pics, code, nm),
    })

spray_metallic_colors = []
for row in raw["Sprayed Metallic Color"][1:]:
    clean = [c for c in row if c is not None]
    if not clean:
        continue
    name = row[0]
    code = row[1] if len(row) > 1 else None
    if not name:
        continue
    nm = name.replace("\n", " ").strip()
    spray_metallic_colors.append({
        "code": code,
        "name": nm,
        "picture": _surface_pic(metal_pics, code, nm),
    })

# ---------------------------------------------------------------------------
# 7. HARDWARE
# ---------------------------------------------------------------------------
hardware_list = []
for row in raw["hardware"][1:]:
    clean = [c for c in row if c is not None]
    if not clean:
        continue
    name = row[0]
    price_raw = row[2] if len(row) > 2 else None
    code = row[3] if len(row) > 3 else None
    color_raw = row[4] if len(row) > 4 else None
    color_code = row[5] if len(row) > 5 else None

    if not name:
        continue

    price = None
    if price_raw:
        m = re.search(r'(\d+)', str(price_raw))
        if m:
            price = int(m.group(1))

    colors = parse_hardware_colors(color_raw) if color_raw else []

    clean_name = name.replace("\n", " ").strip()
    pic = None
    if code:
        pic = hardware_pics.get(str(code).strip())
    else:
        pic = hardware_name_pics.get(_slug(clean_name))

    if code and not pic:
        pic = _existing_catalog_url("hardware", str(code).strip())
    # Blum cup hinge row often has no embedded image; reuse on-disk Blum art or Sensys shot.
    if not pic and code and str(code).strip().upper().replace(" ", "") in ("HG-BLUM", "HGBLUM"):
        for fb in ("HG-BLUM", "Pipe_Hinge_salichi_blum_", "HG-SEN"):
            pic = _existing_catalog_url("hardware", fb)
            if pic:
                break
    if not pic and "blum" in clean_name.lower() and "cup" in clean_name.lower():
        for fb in ("HG-BLUM", "Pipe_Hinge_salichi_blum_", "HG-SEN"):
            pic = _existing_catalog_url("hardware", fb)
            if pic:
                break

    hardware_list.append({
        "code": code,
        "name": clean_name,
        "allowedColors": colors,
        "pricePerPiece": price,
        "picture": pic,
    })

# HG-BLUM row often has empty price cell; mirror Sensys until Excel is filled.
_sen_row = next(
    (x for x in hardware_list if str(x.get("code") or "").strip().upper() == "HG-SEN"),
    None,
)
_sen_price = _sen_row.get("pricePerPiece") if _sen_row else None
for _x in hardware_list:
    if (
        str(_x.get("code") or "").strip().upper() == "HG-BLUM"
        and _x.get("pricePerPiece") is None
        and _sen_price is not None
    ):
        _x["pricePerPiece"] = _sen_price

# Placeholder row when Cabinet Door references a hinge code missing from the hardware sheet
_hw_codes_norm = {
    _hinge_code_fingerprint(x["code"])
    for x in hardware_list
    if x.get("code")
}
if "jm47012" not in _hw_codes_norm:  # JM470-1-2
    hardware_list.append({
        "code": "JM470-1-2",
        "name": "JM470-1-2 hinge (placeholder — add photo/price on hardware sheet)",
        "allowedColors": [],
        "pricePerPiece": None,
        "picture": None,
    })

# ---------------------------------------------------------------------------
# 8. HANDLES — union of Cabinet Door handleOptions + codes on frames + handle sheet pics
# ---------------------------------------------------------------------------
handle_registry = {}


def _reg_handle(code: str, name: str | None, picture: str | None):
    if not code:
        return
    code = str(code).strip()
    if not code:
        return
    if code.lower() == "fixed":
        code = "FIXED"
    if code not in handle_registry:
        handle_registry[code] = {
            "code": code,
            "name": (name or code).replace("\n", " ").strip(),
            "surfaceFinishing": None,
            "allowedColors": ["color swatch colors"],
            "picture": picture,
        }
    else:
        if picture and not handle_registry[code].get("picture"):
            handle_registry[code]["picture"] = picture


for f in frames:
    for opt in f.get("handleOptions") or []:
        _reg_handle(opt["code"], opt.get("name"), opt.get("picture"))
    if f.get("fixedHandleCode"):
        fc = f["fixedHandleCode"]
        _reg_handle(fc, fc, handle_pics.get(fc))
    if f.get("handleWorkflow") == "legacy_catalog" and f.get("matchedHandle"):
        mh_raw = f["matchedHandle"]
        for part in re.split(r"[/\n,，]+", mh_raw):
            part = part.strip()
            if len(part) < 2:
                continue
            pl = part.lower()
            if pl in (
                "separate handle",
                "routed handle(cnc)",
                "cnc",
                "handle",
                "fixed",
                "pull",
            ):
                continue
            if "(" in part:
                for c in _parse_paren_codes(part):
                    _reg_handle(c, c, handle_pics.get(c))
            elif re.match(r"^[A-Za-z0-9][A-Za-z0-9._-]{1,}$", part):
                _reg_handle(part, part, handle_pics.get(part))

if any(f.get("handleWorkflow") == "cnc" for f in frames):
    _reg_handle("CNC", "铣型拉手 / Routed CNC handle", handle_pics.get("CNC"))

# Handle catalog entries come from Cabinet Door rows (per-frame options + fixed codes).
# Optional legacy "handle" sheet images still fill handle_pics for matching codes above.

for _code, _row in list(handle_registry.items()):
    if not _row.get("picture"):
        _p = handle_pics.get(_code)
        if _p:
            _row["picture"] = _p

handle_list = sorted(handle_registry.values(), key=lambda x: str(x["code"]))

if "FIXED" in handle_registry:
    handle_registry["FIXED"]["name"] = "固定款式拉手（位置可选）"
    handle_list = sorted(handle_registry.values(), key=lambda x: str(x["code"]))

if "CNC" in handle_registry:
    handle_registry["CNC"]["name"] = "铣型拉手 / Routed CNC handle"
    handle_list = sorted(handle_registry.values(), key=lambda x: str(x["code"]))


# ---------------------------------------------------------------------------
# GENERATE TypeScript
# ---------------------------------------------------------------------------
def ts_obj(obj, indent=2):
    lines = []
    pad = " " * indent
    for k, v in obj.items():
        if isinstance(v, dict):
            inner = ts_obj(v, indent + 2)
            lines.append(f"{pad}{k}: {inner},")
        elif isinstance(v, list):
            if not v:
                lines.append(f"{pad}{k}: [],")
            elif isinstance(v[0], dict):
                items = ",\n".join(ts_obj(item, indent + 4) for item in v)
                lines.append(f"{pad}{k}: [\n{items}\n{pad}],")
            else:
                inner = ", ".join(to_ts_val(x) for x in v)
                lines.append(f"{pad}{k}: [{inner}],")
        else:
            lines.append(f"{pad}{k}: {to_ts_val(v)},")
    return "{\n" + "\n".join(lines) + "\n" + " " * (indent - 2) + "}"


def ts_array(arr, name, indent=0):
    pad = " " * indent
    if not arr:
        empty_types = {
            "woodVeneerList": "readonly WoodVeneer[]",
        }
        if name in empty_types:
            return f"export const {name} = [] as {empty_types[name]};\n"
    items = []
    for obj in arr:
        items.append(ts_obj(obj, indent + 4))
    joined = ",\n".join(items)
    return f"export const {name} = [\n{joined}\n] as const;\n"


def ts_obj_top(groups, name):
    lines = []
    for key, arr in groups.items():
        items = ",\n".join(ts_obj(obj, 6) for obj in arr)
        lines.append(f"  {key}: [\n{items}\n  ],")
    body = "\n".join(lines)
    return f"export const {name} = {{\n{body}\n}} as const;\n"


# Build output
out_parts = []

# ---- Type definitions ----
out_parts.append("""// =============================================================================
// Gainer Door Configurator — Structured Data
// Auto-generated from source Excel: \"318 door information(3).xlsx\"
// =============================================================================

// ---------------------------------------------------------------------------
// Type Definitions
// ---------------------------------------------------------------------------

export interface SizeLimits {
  minW: number | null;
  maxW: number | null;
  minH: number | null;
  maxH: number | null;
}

export type FrameCategory = 'cabinet' | 'room';
export type FrameMountingType = 'insert' | 'cover';

export type HandleWorkflow =
  | 'none'
  | 'vshape'
  | 'fixed'
  | 'separate'
  | 'cnc'
  | 'legacy_catalog';

export interface FrameHandleOption {
  readonly code: string;
  readonly name: string;
  readonly picture: string | null;
}

export interface Frame {
  id: string;
  code: string;
  frameCategory: FrameCategory;
  doorType: string;
  allowedFillers: readonly string[];
  standardFillers: readonly string[];
  fillerThicknessLimit: readonly number[];
  allowedFinishing: readonly string[];
  specificColors: readonly string[] | null;
  sizeLimits: SizeLimits;
  doorThickness: number | null;
  /** Raw Excel profile cell, e.g. "28mm/Cover". */
  frameProfileLabel: string | null;
  /** Parsed from profile cell: 卡槽 insert / 贴面 cover. */
  mountingType: FrameMountingType | null;
  /** Pull / CNC / none / V-shape / fixed / legacy text match. */
  handleWorkflow: HandleWorkflow;
  /** Per-row handle SKUs from Excel handle-picture columns (Separate handle). */
  handleOptions: readonly FrameHandleOption[];
  /** Preset code for V-shape / fixed (no tile pick). */
  fixedHandleCode: string | null;
  /** Door diagram for handle position UI; falls back to main photo. */
  handleDiagramPicture: string | null;
  /** Legacy: raw Excel handle cell for catalog matching. */
  matchedHandle: string | null;
  matchedHardware: string | null;
  /** Hinge product codes from Excel (e.g. AIRHINGE / HG-BLUM), matched to hardwareList. */
  hingeCodes: readonly string[];
  hardwareColors: readonly string[];
  /** Column D — side view; opened from button on frame card. */
  pictureSideView: string | null;
  /** Column E — profile section; opened from button on frame card. */
  pictureProfile: string | null;
  /** Column R — hinge photo from sheet (Step 6). */
  hingePicture: string | null;
  /** [main, side, profile] with nulls removed (compatibility). */
  pictures: readonly string[];
  /** Column C — main list thumbnail; if C is empty, falls back to D then E. */
  picture: string | null;
}

export interface FrameGlassPricing {
  normalGlass: number | null;
  blackGlass: number | null;
  coatedGlass: number | null;
}

export interface Glass {
  code: string;
  name: string;
  type: 'glass';
  craft: string | null;
  craftCode: string | null;
  thicknesses: readonly number[];
  silkScreen: boolean;
  pricingType: 'standard' | 'custom';
  priceSqm: number | null;
  picture: string | null;
}

export interface Leather {
  code: string;
  name: string;
  type: 'leather';
  thicknesses: readonly number[];
  baseMaterial: 'Honeycomb Aluminum';
  pricingType: 'custom';
  picture: string | null;
}

export interface WoodVeneer {
  code: string;
  name: string;
  type: 'woodVeneer';
  thicknesses: readonly number[];
  pricingType: 'custom';
  picture: string | null;
}

export interface QuartzStone {
  code: string;
  name: string;
  type: 'quartzStone';
  surface: string | null;
  thicknesses: readonly number[];
  pricingType: 'custom';
  picture: string | null;
}

export type Filler = Glass | Leather | WoodVeneer | QuartzStone;

export interface SurfaceColor {
  code: string | null;
  name: string;
  picture: string | null;
}

export interface SurfaceFinishes {
  anodize: SurfaceColor[];
  spraySoftTouch: SurfaceColor[];
  sprayMetallic: SurfaceColor[];
}

export interface Hardware {
  code: string | null;
  name: string;
  allowedColors: readonly string[];
  pricePerPiece: number | null;
  picture: string | null;
}

export interface Handle {
  code: string;
  name: string;
  surfaceFinishing: string | null;
  allowedColors: readonly string[];
  picture: string | null;
}
""")

# ---- Frames ----
out_parts.append("\n// ---------------------------------------------------------------------------")
out_parts.append("// 1. Frames — 柜门 cabinet / 房门 room (same Excel sheet, split by section)")
out_parts.append("// ---------------------------------------------------------------------------\n")
out_parts.append(ts_array(frames, "frames"))

# ---- Standard glass combo pricing (parsed from Excel price column) ----
def _ts_frame_pricing_map(prices: dict) -> str:
    lines = []
    for k in sorted(prices.keys()):
        v = prices[k]
        inner = ", ".join(
            f"{x}: {to_ts_val(v[x])}" for x in ("normalGlass", "blackGlass", "coatedGlass")
        )
        lines.append(f"  {to_ts_val(k)}: {{ {inner} }}")
    return "export const frameStandardGlassPricingByCode: Record<string, FrameGlassPricing> = {\n" + ",\n".join(lines) + "\n};\n"

out_parts.append("\n// ---------------------------------------------------------------------------")
out_parts.append("// 1b. Frame + standard glass (G01/G33/G36 tier) reference pricing from Excel")
out_parts.append("// ---------------------------------------------------------------------------\n")
out_parts.append(_ts_frame_pricing_map(frame_standard_glass_prices))

# ---- Glass ----
out_parts.append("\n// ---------------------------------------------------------------------------")
out_parts.append("// 2. Glass Fillers")
out_parts.append("// ---------------------------------------------------------------------------\n")
out_parts.append(ts_array(glass_list, "glassList"))

# ---- Leather ----
out_parts.append("\n// ---------------------------------------------------------------------------")
out_parts.append("// 3. Leather Fillers (baseMaterial: Honeycomb Aluminum)")
out_parts.append("// ---------------------------------------------------------------------------\n")
out_parts.append(ts_array(leather_list, "leatherList"))

# ---- Wood Veneer ----
out_parts.append("\n// ---------------------------------------------------------------------------")
out_parts.append("// 4. Wood Veneer Fillers")
out_parts.append("// ---------------------------------------------------------------------------\n")
out_parts.append(ts_array(wood_list, "woodVeneerList"))

# ---- Quartz Stone ----
out_parts.append("\n// ---------------------------------------------------------------------------")
out_parts.append("// 5. Quartz Stone Fillers")
out_parts.append("// ---------------------------------------------------------------------------\n")
out_parts.append(ts_array(quartz_list, "quartzStoneList"))

# ---- Surface Finishes ----
out_parts.append("\n// ---------------------------------------------------------------------------")
out_parts.append("// 6. Surface Finishes")
out_parts.append("// ---------------------------------------------------------------------------\n")
out_parts.append(ts_obj_top({
    "anodize": anodize_colors,
    "spraySoftTouch": spray_soft_colors,
    "sprayMetallic": spray_metallic_colors,
}, "surfaceFinishes"))

# ---- Hardware ----
out_parts.append("\n// ---------------------------------------------------------------------------")
out_parts.append("// 7. Hardware (Hinges & Accessories)")
out_parts.append("// ---------------------------------------------------------------------------\n")
out_parts.append(ts_array(hardware_list, "hardwareList"))

out_parts.append("""
function _hardwareCodeFingerprint(s: string): string {
  return String(s)
    .trim()
    .toLowerCase()
    .replace(/[\\s\\-_]+/g, '');
}

export function getHardwareByCode(code: string | null | undefined): Hardware | null {
  if (!code) return null;
  const raw = String(code).trim().toLowerCase();
  const variants = [raw, raw.replace(/^hg-/, ''), `hg-${raw}`];
  for (const c of variants) {
    const h = hardwareList.find(
      (x) => x.code != null && String(x.code).trim().toLowerCase() === c,
    );
    if (h) return h;
  }
  const fp = _hardwareCodeFingerprint(String(code));
  if (fp.length > 0) {
    const h = hardwareList.find(
      (x) => x.code != null && _hardwareCodeFingerprint(String(x.code)) === fp,
    );
    if (h) return h;
  }
  return null;
}
""")

# ---- Handles ----
out_parts.append("\n// ---------------------------------------------------------------------------")
out_parts.append("// 8. Handles")
out_parts.append("// ---------------------------------------------------------------------------\n")
out_parts.append(ts_array(handle_list, "handleList"))

# ---- Aggregate filler map ----
out_parts.append("""
// ---------------------------------------------------------------------------
// 9. Filler Lookup Map (for state machine filtering)
// ---------------------------------------------------------------------------

export const fillerMap = {
  glass: glassList,
  leather: leatherList,
  woodVeneer: woodVeneerList,
  quartzStone: quartzStoneList,
} as const;
""")

# ---- State-machine filter helpers ----
out_parts.append("""
// ---------------------------------------------------------------------------
// 10. State-Machine Filtering Logic
// ---------------------------------------------------------------------------

export function getAvailableFrames(w: number, h: number): (Frame & { disabled: boolean })[] {
  return frames.map((frame) => {
    const { minW, maxW, minH, maxH } = frame.sizeLimits;
    const disabled =
      (minW !== null && w < minW) ||
      (maxW !== null && w > maxW) ||
      (minH !== null && h < minH) ||
      (maxH !== null && h > maxH);
    return { ...frame, disabled };
  });
}

export function getAvailableFinishes(frame: Frame): SurfaceFinishes {
  const allowed = frame.allowedFinishing;
  return {
    anodize: allowed.includes('anodize') ? [...surfaceFinishes.anodize] : [],
    spraySoftTouch: allowed.includes('spraySoftTouch') ? [...surfaceFinishes.spraySoftTouch] : [],
    sprayMetallic: allowed.includes('sprayMetallic') ? [...surfaceFinishes.sprayMetallic] : [],
  };
}

export function getAvailableFillers(frame: Frame) {
  const allowedTypes = frame.allowedFillers;
  const thicknessLimit = frame.fillerThicknessLimit;

  const result: Filler[] = [];

  if (allowedTypes.includes('glass')) {
    for (const g of glassList) {
      const validThicknesses = thicknessLimit.length > 0
        ? g.thicknesses.filter((t) => thicknessLimit.includes(t))
        : g.thicknesses;
      if (validThicknesses.length > 0) {
        result.push({ ...g, thicknesses: validThicknesses });
      }
    }
  }

  if (allowedTypes.includes('leather')) {
    result.push(...leatherList);
  }

  if (allowedTypes.includes('wood veneer') || allowedTypes.includes('woodVeneer')) {
    result.push(...woodVeneerList);
  }

  if (allowedTypes.includes('quartz stone') || allowedTypes.includes('quartzStone')) {
    result.push(...quartzStoneList);
  }

  return result;
}

export function getMatchedHandles(frame: Frame): Handle[] {
  if (frame.handleOptions.length > 0) {
    return frame.handleOptions.map((o) => ({
      code: o.code,
      name: o.name,
      surfaceFinishing: null,
      allowedColors: ['color swatch colors'],
      picture: o.picture,
    })) as Handle[];
  }
  if (!frame.matchedHandle) return [];
  const raw = frame.matchedHandle;
  const codes = raw.split(/[\\n\\/]/).map((s) => s.trim().toLowerCase()).filter(Boolean);
  return handleList.filter(
    (h) => codes.some((c) => h.code.toLowerCase() === c || h.name.toLowerCase().includes(c))
  );
}

export function getMatchedHandle(frame: Frame): Handle | null {
  const list = getMatchedHandles(frame);
  if (list.length > 0) return list[0];
  if (frame.fixedHandleCode) {
    const hit = handleList.find((h) => h.code === frame.fixedHandleCode);
    if (hit) return hit;
    return {
      code: frame.fixedHandleCode,
      name: frame.fixedHandleCode,
      surfaceFinishing: null,
      allowedColors: ['color swatch colors'],
      picture: null,
    } as Handle;
  }
  return null;
}

export function calculateHingeQty(heightMm: number): { qty: number; usePivot: boolean } {
  if (heightMm <= 2000) return { qty: 2, usePivot: false };
  if (heightMm <= 2500) return { qty: 4, usePivot: false };
  return { qty: 0, usePivot: true };
}
""")

# Write final file
output = "\n".join(out_parts)
with open("src/data.ts", "w", encoding="utf-8") as f:
    f.write(output)

print(f"Generated data.ts")
print(f"  Frames: {len(frames)}")
print(f"  Glass:  {len(glass_list)}")
print(f"  Leather: {len(leather_list)}")
print(f"  Wood Veneer: {len(wood_list)}")
print(f"  Quartz Stone: {len(quartz_list)}")
print(f"  Anodized Colors: {len(anodize_colors)}")
print(f"  Spray Soft-Touch: {len(spray_soft_colors)}")
print(f"  Spray Metallic: {len(spray_metallic_colors)}")
print(f"  Hardware: {len(hardware_list)}")
print(f"  Handles: {len(handle_list)}")
